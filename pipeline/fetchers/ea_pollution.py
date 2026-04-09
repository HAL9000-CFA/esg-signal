"""
UK Environment Agency Pollution Inventory fetcher.

Data covers 2013–2024 (published annually, last updated Nov 2025).
Source: https://www.data.gov.uk/dataset/cfd94301-a2f2-48a2-9915-e477ca6d8b7e/pollution-inventory

File format
-----------
Each year is a ZIP containing an XLSX with three sheets:
  '{YEAR} Substances'        — facility-level emissions to air/land/water
  '{YEAR} Waste Transfers'   — off-site waste disposal/recovery (incl. hazardous)
  '{YEAR} Radioactive Wastes'— radioactive material releases

All sheets have a 9-row preamble; headers are on row 10, data from row 11.
OPERATOR NAME is column index 2 in Substances and Waste Transfers,
and column index 1 in Radioactive Wastes.

Name matching
-------------
The EA dataset uses legal subsidiary names (e.g. 'BP Oil UK Limited') rather
than parent company names ('BP P.L.C.').  This fetcher extracts the first
meaningful token from the supplied company_name and applies a word-boundary
regex to avoid false positives (e.g. 'ABP' matching 'BP').

Caching
-------
Downloaded XLSX files are cached at data/raw/ea_pollution/{year}.xlsx.
A file is re-downloaded only if absent.  The CKAN API is queried to discover
the current year's ZIP URL; the package UUID is stable.
"""

import io
import logging
import re
import zipfile
from pathlib import Path
from typing import Optional

import pandas as pd
import requests

from pipeline.fetchers.base_regulatory import BaseRegulatoryFetcher

LOGGER = logging.getLogger(__name__)

_CKAN_API = "https://www.data.gov.uk/api/3/action/package_show"
_PACKAGE_UUID = "cfd94301-a2f2-48a2-9915-e477ca6d8b7e"
_CACHE_DIR = Path("data/raw/ea_pollution")

# Row index (0-based) of the header row in every sheet
_HEADER_ROW = 9  # row 10 in 1-based Excel terms

# OPERATOR NAME column index (0-based) per sheet family
_OP_COL = {
    "substances": 2,
    "waste": 2,
    "radioactive": 1,
}

# Legal suffixes stripped when extracting the core search token
_LEGAL_SUFFIXES = re.compile(
    r"\b(P\.?L\.?C\.?|LIMITED|LTD|INC|CORP|LLC|GROUP|HOLDINGS?|UK|"
    r"INTERNATIONAL|GLOBAL|CO|COMPANY|PLC)\b",
    re.IGNORECASE,
)


def _search_tokens(company_name: str) -> list[str]:
    """
    Extract up to two meaningful tokens from a company name for EA dataset matching.

    Strips legal suffixes, punctuation, and single characters, then returns the
    first 1–2 meaningful words.  Two tokens are preferred because single generic
    words ('British', 'National', 'Anglo') over-match unrelated companies.

    Examples:
      'BP P.L.C.'            → ['BP']
      'Shell PLC'            → ['Shell']
      'National Grid PLC'    → ['National', 'Grid']
      'Anglo American PLC'   → ['Anglo', 'American']
      'AstraZeneca PLC'      → ['AstraZeneca']
      'Legal & General PLC'  → ['Legal', 'General']
    """
    cleaned = re.sub(r"[.()\[\]&,]", " ", company_name)
    cleaned = _LEGAL_SUFFIXES.sub(" ", cleaned)
    # Drop single characters (e.g. the 'a' left from '& a')
    tokens = [t for t in cleaned.split() if len(t) > 1]
    return tokens[:2] if len(tokens) >= 2 else tokens[:1]


def _build_pattern(company_name: str) -> re.Pattern:
    """
    Build a word-boundary regex from the company name tokens.

    Two tokens -> phrase match: r'\\bToken1\\s+Token2\\b'
    One token  -> single match: r'\\bToken\\b'
    """
    tokens = _search_tokens(company_name)
    if len(tokens) == 2:
        expr = r"\b" + re.escape(tokens[0]) + r"\s+" + re.escape(tokens[1]) + r"\b"
    elif tokens:
        expr = r"\b" + re.escape(tokens[0]) + r"\b"
    else:
        expr = re.escape(company_name.split()[0])
    return re.compile(expr, re.IGNORECASE)


def _get_zip_url(year: int) -> Optional[str]:
    """Query the CKAN API and return the ZIP download URL for the given year."""
    try:
        r = requests.get(_CKAN_API, params={"id": _PACKAGE_UUID}, timeout=15)
        r.raise_for_status()
        resources = r.json().get("result", {}).get("resources", [])
    except Exception as exc:
        LOGGER.warning("EA Pollution: CKAN API failed: %s", exc)
        return None

    # Prefer resources whose name mentions the target year
    for res in resources:
        name = res.get("name", "") or ""
        url = res.get("url", "") or ""
        fmt = (res.get("format", "") or "").upper()
        if str(year) in name and fmt in ("ZIP", "XLSX", "XLS", ""):
            return url

    # Fallback: any resource mentioning the year
    for res in resources:
        url = res.get("url", "") or ""
        if str(year) in url:
            return url

    return None


def _download_xlsx(year: int) -> Optional[Path]:
    """
    Return the local path to the cached XLSX for the given year.
    Downloads and extracts from the ZIP if not already cached.
    """
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cached = _CACHE_DIR / f"{year}.xlsx"
    if cached.exists():
        LOGGER.info("EA Pollution: using cached XLSX for year=%d at %s", year, cached)
        return cached

    url = _get_zip_url(year)
    if not url:
        LOGGER.warning("EA Pollution: could not discover ZIP URL for year=%d", year)
        return None

    LOGGER.info("EA Pollution: downloading year=%d from %s", year, url)
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
    except requests.exceptions.Timeout:
        LOGGER.warning("EA Pollution: download timed out (120s) for year=%d", year)
        return None
    except Exception as exc:
        LOGGER.warning("EA Pollution: download failed for year=%d: %s", year, exc)
        return None

    content = r.content

    # If the response is already an XLSX (some years are served directly)
    if content[:4] == b"PK\x03\x04":
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
                if not xlsx_names:
                    LOGGER.warning("EA Pollution: ZIP for year=%d contains no XLSX", year)
                    return None
                # Pick the Pollution Inventory XLSX (not the briefing PDF)
                xlsx_name = next(
                    (n for n in xlsx_names if "pollution inventory" in n.lower()),
                    xlsx_names[0],
                )
                with zf.open(xlsx_name) as f:
                    cached.write_bytes(f.read())
                LOGGER.info("EA Pollution: extracted %s → %s", xlsx_name, cached)
                return cached
        except zipfile.BadZipFile:
            pass

    # Try saving as XLSX directly
    try:
        cached.write_bytes(content)
        LOGGER.info("EA Pollution: saved XLSX directly for year=%d", year)
        return cached
    except Exception as exc:
        LOGGER.warning("EA Pollution: could not save XLSX for year=%d: %s", year, exc)
        return None


# Normalise each sheet's varying column names to a common schema.
# Each sheet has different names for the same concepts; normalisation
# lets us concat all three into a single readable DataFrame.
_NORMALISE = {
    # operator identity (common to all sheets)
    "OPERATOR NAME": "operator_name",
    "SITE ADDRESS": "site_address",
    "SITE POSTCODE": "postcode",
    "EA AREA NAME": "ea_area",
    "AUTHORISATION ID / PERMIT ID": "permit_id",
    "AUTHORISATION ID /PERMIT ID": "permit_id",  # Radioactive has no space before /
    "ACTIVITY DESCRIPTION": "activity_description",
    "REGULATED INDUSTRY SECTOR": "sector",
    "REGULATED INDUSTRY SUB SECTOR": "sub_sector",
    # substance / waste description
    "SUBSTANCE NAME": "substance_or_waste",  # Substances + Radioactive
    "EWC DESCRIPTION": "substance_or_waste",  # Waste Transfers
    "EWC ID": "ewc_id",
    # quantity — different names and units per sheet
    "QUANTITY RELEASED (kg)": "quantity",  # Substances (kg)
    "QUANTITY RELEASED (Tonnes)": "quantity",  # Waste Transfers (tonnes)
    "QUANTITY RELEASED": "quantity",  # Radioactive (unit varies)
    "REPORTING THRESHOLD (kg)": "threshold",
    "REPORTING THRESHOLD (UNIT OF MEASURE)": "threshold",
    "ANNUAL REPORTING THRESHOLD": "threshold",
    "UNIT OF MEASURE": "unit",
    # route / disposal
    "ROUTE NAME": "route",
    "ROUTE DESCRIPTION": "route_description",
    # waste-specific
    "HAZARDOUS WASTE": "hazardous",
    "RELEASE LEVEL": "release_level",
}


def _parse_sheet(ws, op_col: int, pattern: re.Pattern) -> pd.DataFrame:
    """
    Parse one openpyxl worksheet, filter rows where OPERATOR NAME matches
    pattern, normalise column names, and return as a DataFrame.
    """
    headers_raw = None
    rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == _HEADER_ROW:
            headers_raw = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
            continue
        if i < _HEADER_ROW:
            continue
        if headers_raw is None:
            continue
        op = row[op_col] if len(row) > op_col else None
        if op and pattern.search(str(op)):
            rows.append(dict(zip(headers_raw, row)))

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # Capture the quantity unit BEFORE renaming so it isn't lost.
    # Each sheet uses a different column name that encodes the unit:
    #   Substances      → QUANTITY RELEASED (kg)      → unit = "kg"
    #   Waste Transfers → QUANTITY RELEASED (Tonnes)  → unit = "tonnes"
    #   Radioactive     → QUANTITY RELEASED + UNIT OF MEASURE column
    if "QUANTITY RELEASED (kg)" in df.columns:
        df["quantity_unit"] = "kg"
    elif "QUANTITY RELEASED (Tonnes)" in df.columns:
        df["quantity_unit"] = "tonnes"
    elif "QUANTITY RELEASED" in df.columns and "UNIT OF MEASURE" in df.columns:
        df["quantity_unit"] = df["UNIT OF MEASURE"]
    else:
        df["quantity_unit"] = None

    # Rename to normalised schema, keep any columns not in the map as-is
    df = df.rename(columns={k: v for k, v in _NORMALISE.items() if k in df.columns})
    return df


class EAPollutionFetcher(BaseRegulatoryFetcher):
    def fetch(self, company_name: str, year: Optional[int] = None) -> pd.DataFrame:
        """
        Fetch EA Pollution Inventory records for the given company.

        Searches all three sheets (Substances, Waste Transfers, Radioactive Wastes)
        and returns a combined DataFrame with a 'data_type' column indicating
        the source sheet.

        Args:
            company_name: Company name as used in the pipeline (e.g. 'BP P.L.C.').
                          The core token is extracted automatically for EA matching.
            year:         Year to fetch (defaults to 2024, the most recent available).

        Returns:
            DataFrame with columns from all matched sheets, plus 'data_type'.
            Empty DataFrame if no records found.
        """
        target_year = year or 2024
        tokens = _search_tokens(company_name)
        primary_pattern = _build_pattern(company_name)

        LOGGER.info(
            "EA Pollution: fetching year=%d for company=%r (tokens=%s pattern=%r)",
            target_year,
            company_name,
            tokens,
            primary_pattern.pattern,
        )

        xlsx_path = _download_xlsx(target_year)
        if xlsx_path is None:
            LOGGER.warning("EA Pollution: no XLSX available for year=%d", target_year)
            return pd.DataFrame()

        try:
            import openpyxl

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as exc:
            LOGGER.warning("EA Pollution: could not open XLSX %s: %s", xlsx_path, exc)
            return pd.DataFrame()

        sheet_configs = [
            (f"{target_year} Substances", "substances", _OP_COL["substances"]),
            (f"{target_year} Waste Transfers", "waste_transfers", _OP_COL["waste"]),
            (f"{target_year} Radioactive Wastes", "radioactive_wastes", _OP_COL["radioactive"]),
        ]

        frames = []
        for sheet_name, data_type, op_col in sheet_configs:
            if sheet_name not in wb.sheetnames:
                LOGGER.info("EA Pollution: sheet %r not found in year=%d", sheet_name, target_year)
                continue
            ws = wb[sheet_name]

            df = _parse_sheet(ws, op_col, primary_pattern)

            if not df.empty:
                df["data_type"] = data_type
                df["year"] = target_year
                frames.append(df)
                LOGGER.info(
                    "EA Pollution: %d rows from %r (pattern=%r)",
                    len(df),
                    sheet_name,
                    primary_pattern.pattern,
                )
            else:
                LOGGER.info("EA Pollution: no rows in %r for company=%r", sheet_name, company_name)

        wb.close()

        if not frames:
            LOGGER.info(
                "EA Pollution: no records found for company=%r (token=%r) year=%d",
                company_name,
                tokens,
                target_year,
            )
            return pd.DataFrame()

        result = pd.concat(frames, ignore_index=True)
        LOGGER.info(
            "EA Pollution: done — %d total records for company=%r year=%d",
            len(result),
            company_name,
            target_year,
        )
        return result

    def save(self, df: pd.DataFrame, raw_path: str, processed_path: str) -> None:
        Path(raw_path).parent.mkdir(parents=True, exist_ok=True)
        Path(processed_path).parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(raw_path, index=False)
        df.to_csv(processed_path, index=False)

"""
ESG Signal — Streamlit briefing interface (issue #16).

Page flow
---------
1. Input form: ticker, company name, index, optional DCF upload.
2. "Run Analysis" triggers the Airflow DAG via the REST API and stores the
   dag_run_id in session state.
3. Polling loop: the page auto-refreshes every 5 s until the run finishes.
4. Results: CompanyProfile metadata, CredibilityReport (overall + per-factor
   panels), optional DCF scenario table.
5. Export: PDF (ReportLab) and JSON download buttons.

DCF upload
----------
The uploaded .xlsx is saved to data/dcf_uploads/ (shared Docker volume).
The file is parsed locally with openpyxl to show a preview (sheet names and
label count) before the analysis runs.  The Airflow DAG receives the
Airflow-side absolute path to the same file.

Environment variables (all optional — defaults work inside Docker Compose)
--------------------------------------------------------------------------
AIRFLOW_URL          Base URL of the Airflow webserver  (default: http://localhost:8080)
AIRFLOW_USER         Airflow basic-auth username         (default: admin)
AIRFLOW_PASSWORD     Airflow basic-auth password         (default: admin)
DCF_UPLOAD_LOCAL_DIR Where Streamlit writes DCF uploads (default: data/dcf_uploads)
DCF_UPLOAD_AIRFLOW_DIR  Airflow-side path for the same dir
                        (default: /opt/airflow/data/dcf_uploads)
"""

import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
import streamlit as st

from pipeline.models import (
    CommitmentCheck,
    CompanyProfile,
    CredibilityReport,
    DcfLineItem,
    DcfMapperResult,
    EsgDcfMapping,
    FactorScore,
    FilingMetadata,
)
from ui.components import (
    credibility_header,
    dcf_mapping_panel,
    factor_panel,
    source_citation,
)
from ui.export import to_json, to_pdf

LOGGER = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_AIRFLOW_URL = os.getenv("AIRFLOW_URL", "http://localhost:8080").rstrip("/")
_AIRFLOW_USER = os.getenv("AIRFLOW_USER", "admin")
_AIRFLOW_PASSWORD = os.getenv("AIRFLOW_PASSWORD", "admin")
_AIRFLOW_AUTH = (_AIRFLOW_USER, _AIRFLOW_PASSWORD)
_AIRFLOW_API = f"{_AIRFLOW_URL}/api/v1"
_DAG_ID = "esg_signal"

_DCF_LOCAL_DIR = Path(os.getenv("DCF_UPLOAD_LOCAL_DIR", "data/dcf_uploads"))
_DCF_AIRFLOW_DIR = os.getenv("DCF_UPLOAD_AIRFLOW_DIR", "/opt/airflow/data/dcf_uploads")

_POLL_INTERVAL_SECONDS = 5

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="ESG Signal",
    page_icon="🟢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ---------------------------------------------------------------------------
# Session state helpers
# ---------------------------------------------------------------------------


def _init_state() -> None:
    defaults = {
        "run_state": "idle",  # idle | running | success | failed
        "dag_run_id": None,
        "profile_dict": None,
        "credibility_dict": None,
        "dcf_dict": None,
        "error_message": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _reset_state() -> None:
    for k in (
        "run_state",
        "dag_run_id",
        "profile_dict",
        "credibility_dict",
        "dcf_dict",
        "error_message",
    ):
        if k in st.session_state:
            del st.session_state[k]
    _init_state()


# ---------------------------------------------------------------------------
# Airflow REST API helpers
# ---------------------------------------------------------------------------


def _trigger_dag(ticker: str, company_name: str, index: str, dcf_airflow_path: str) -> str:
    """
    Trigger the esg_signal DAG and return the dag_run_id.

    Raises requests.HTTPError on API failure.
    """
    run_id = f"ui__{ticker.upper()}__{datetime.utcnow().strftime('%Y%m%dT%H%M%S')}"
    payload = {
        "dag_run_id": run_id,
        "conf": {
            "ticker": ticker,
            "company_name": company_name,
            "index": index,
            "dcf_path": dcf_airflow_path,
        },
    }
    resp = requests.post(
        f"{_AIRFLOW_API}/dags/{_DAG_ID}/dagRuns",
        json=payload,
        auth=_AIRFLOW_AUTH,
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["dag_run_id"]


def _get_dag_run_state(run_id: str) -> str:
    """Return the Airflow dag run state string (e.g. 'running', 'success', 'failed')."""
    resp = requests.get(
        f"{_AIRFLOW_API}/dags/{_DAG_ID}/dagRuns/{run_id}",
        auth=_AIRFLOW_AUTH,
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json().get("state", "unknown")


def _get_xcom(run_id: str, task_id: str) -> Optional[dict]:
    """
    Fetch the return_value XCom for a task instance.

    Returns the decoded dict, or None if unavailable.
    The Airflow REST API returns XCom values as a JSON-encoded string inside
    the "value" field.
    """
    try:
        resp = requests.get(
            f"{_AIRFLOW_API}/dags/{_DAG_ID}/dagRuns/{run_id}"
            f"/taskInstances/{task_id}/xcomEntries/return_value",
            auth=_AIRFLOW_AUTH,
            timeout=15,
        )
        resp.raise_for_status()
        raw = resp.json().get("value")
        if raw is None:
            return None
        # Airflow serialises the Python value to JSON; the REST API wraps it
        # as a JSON string in the "value" field → double-decode.
        return json.loads(raw) if isinstance(raw, str) else raw
    except Exception as exc:
        LOGGER.warning("_get_xcom(%s, %s) failed: %s", run_id, task_id, exc)
        return None


# ---------------------------------------------------------------------------
# From-dict reconstructors (mirrors pipeline/esg_signal_dag.py — no airflow import)
# ---------------------------------------------------------------------------


def _profile_from_dict(d: dict) -> CompanyProfile:
    fm = d.get("latest_annual_filing")
    return CompanyProfile(
        ticker=d.get("ticker"),
        name=d["name"],
        index=d.get("index"),
        sic_code=d.get("sic_code"),
        sic_description=d.get("sic_description"),
        country=d.get("country", ""),
        latest_annual_filing=FilingMetadata(**fm) if fm else None,
        annual_report_text=d.get("annual_report_text"),
        raw_financials=d.get("raw_financials") or {},
        source_urls=d.get("source_urls") or [],
        errors=d.get("errors") or [],
        identifier=d.get("identifier"),
    )


def _credibility_from_dict(d: dict) -> CredibilityReport:
    def _wm(raw):
        return [CommitmentCheck(**c) for c in raw] if raw else None

    factor_scores = [
        FactorScore(
            factor_id=fs["factor_id"],
            factor_name=fs["factor_name"],
            score=fs["score"],
            flag=fs["flag"],
            stream_scores=fs.get("stream_scores") or {},
            evidence=fs.get("evidence") or [],
            sources=fs.get("sources") or [],
            narrative=fs.get("narrative") or "",
            words_money_checks=_wm(fs.get("words_money_checks")),
        )
        for fs in (d.get("factor_scores") or [])
    ]
    return CredibilityReport(
        ticker=d.get("ticker"),
        company_name=d["company_name"],
        sasb_industry=d.get("sasb_industry"),
        factor_scores=factor_scores,
        overall_score=d["overall_score"],
        overall_flag=d["overall_flag"],
        errors=d.get("errors") or [],
    )


def _dcf_from_dict(d: dict) -> Optional[DcfMapperResult]:
    if not d or d.get("skipped"):
        return None

    mappings = [
        EsgDcfMapping(
            factor_id=m["factor_id"],
            factor_name=m["factor_name"],
            financial_impacts=m.get("financial_impacts") or [],
            mapped_line_items=[DcfLineItem(**li) for li in (m.get("mapped_line_items") or [])],
            scenario_low=m.get("scenario_low"),
            scenario_mid=m.get("scenario_mid"),
            scenario_high=m.get("scenario_high"),
            scenario_currency=m.get("scenario_currency", "USD"),
            scenario_source=m.get("scenario_source", ""),
            credibility_flag=m.get("credibility_flag"),
        )
        for m in (d.get("mappings") or [])
    ]
    return DcfMapperResult(
        ticker=d.get("ticker"),
        excel_path=d.get("excel_path", ""),
        sheet_names=d.get("sheet_names") or [],
        line_item_count=d.get("line_item_count", 0),
        mappings=mappings,
        unmapped_factors=d.get("unmapped_factors") or [],
        errors=d.get("errors") or [],
    )


# ---------------------------------------------------------------------------
# DCF file handling
# ---------------------------------------------------------------------------


def _save_dcf_upload(uploaded_file, ticker: str) -> tuple[str, str]:
    """
    Save an uploaded DCF file to the shared data/dcf_uploads/ directory.

    Returns:
        (local_path, airflow_path) — paths as seen by Streamlit and Airflow.
    """
    _DCF_LOCAL_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    filename = f"{ticker.upper()}_{timestamp}.xlsx"
    local_path = _DCF_LOCAL_DIR / filename
    local_path.write_bytes(uploaded_file.getvalue())
    airflow_path = f"{_DCF_AIRFLOW_DIR}/{filename}"
    return str(local_path), airflow_path


def _preview_dcf_upload(uploaded_file) -> tuple[list[str], int]:
    """
    Parse sheet names and label count from an uploaded DCF file.

    Returns:
        (sheet_names, label_count)
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filename=uploaded_file, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        label_count = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=300, max_col=3, values_only=True):
                val = row[0]
                if isinstance(val, str) and len(val.strip()) >= 3:
                    label_count += 1
        wb.close()
        return sheet_names, label_count
    except Exception as exc:
        LOGGER.warning("_preview_dcf_upload: %s", exc)
        return [], 0


# ---------------------------------------------------------------------------
# Main app
# ---------------------------------------------------------------------------


def main() -> None:
    _init_state()

    st.title("ESG Signal")
    st.caption(
        "AI-powered ESG Commitment Credibility Briefing — HAL 9000 Team, University of East Anglia"
    )
    st.markdown("---")

    # ------------------------------------------------------------------
    # Input form (shown when idle or failed)
    # ------------------------------------------------------------------
    if st.session_state.run_state in ("idle", "failed"):
        if st.session_state.run_state == "failed" and st.session_state.error_message:
            st.error(f"Previous run failed: {st.session_state.error_message}")

        with st.form("run_form"):
            st.subheader("Company")
            col1, col2, col3 = st.columns([2, 3, 1])
            with col1:
                ticker = st.text_input(
                    "Ticker",
                    placeholder="e.g. BP or AAPL",
                    help="FTSE 100 or S&P 500 ticker symbol",
                )
            with col2:
                company_name = st.text_input(
                    "Company name *",
                    placeholder="e.g. Amazon.com, Inc.",
                    help=(
                        "Required — used for EDGAR / Companies House filing search "
                        "and job-posting queries. Use the legal name as registered "
                        "(e.g. 'Amazon.com, Inc.' not 'Amazon')."
                    ),
                )
            with col3:
                index = st.selectbox("Index", ["SP500", "FTSE100"])

            st.subheader("DCF Model (optional)")
            dcf_file = st.file_uploader(
                "Upload analyst DCF (.xlsx)",
                type=["xlsx"],
                help=(
                    "Your Excel DCF model.  Only row labels are sent to the AI — "
                    "financial figures never leave your machine."
                ),
            )
            if dcf_file:
                sheet_names, label_count = _preview_dcf_upload(dcf_file)
                if sheet_names:
                    st.success(
                        f"Parsed: {len(sheet_names)} sheet(s) "
                        f"({', '.join(sheet_names[:4])}"
                        + (" …" if len(sheet_names) > 4 else "")
                        + f"), {label_count} row labels"
                    )

            submitted = st.form_submit_button("Run Analysis", type="primary")

        if submitted:
            if not ticker.strip():
                st.error("Ticker is required.")
                st.stop()
            if not company_name.strip():
                st.error("Company name is required.")
                st.stop()

            dcf_airflow_path = ""
            if dcf_file:
                _, dcf_airflow_path = _save_dcf_upload(dcf_file, ticker.strip())

            try:
                run_id = _trigger_dag(
                    ticker=ticker.strip(),
                    company_name=company_name.strip(),
                    index=index,
                    dcf_airflow_path=dcf_airflow_path,
                )
                st.session_state.dag_run_id = run_id
                st.session_state.run_state = "running"
            except Exception as exc:
                st.session_state.run_state = "failed"
                st.session_state.error_message = str(exc)
                st.error(f"Failed to trigger DAG: {exc}")
                st.stop()

            # st.rerun() raises a Streamlit-internal exception to interrupt
            # execution — must be outside try/except to avoid being caught.
            st.rerun()

    # ------------------------------------------------------------------
    # Polling display
    # ------------------------------------------------------------------
    elif st.session_state.run_state == "running":
        run_id = st.session_state.dag_run_id
        st.info(f"Analysis running — DAG run: `{run_id}`")

        with st.spinner("Waiting for pipeline to complete…"):
            try:
                state = _get_dag_run_state(run_id)
            except Exception as exc:
                st.warning(f"Could not reach Airflow: {exc}")
                state = "unknown"

        if state == "success":
            # Fetch XComs
            fetch_xcom = _get_xcom(run_id, "fetch_data")
            credibility_xcom = _get_xcom(run_id, "run_credibility_scorer")
            dcf_xcom = _get_xcom(run_id, "run_dcf_mapper")

            if not credibility_xcom:
                st.session_state.run_state = "failed"
                st.session_state.error_message = "Credibility scorer XCom not found."
                st.rerun()

            st.session_state.profile_dict = fetch_xcom.get("profile") if fetch_xcom else None
            st.session_state.credibility_dict = credibility_xcom
            st.session_state.dcf_dict = dcf_xcom
            st.session_state.run_state = "success"
            st.rerun()

        elif state in ("failed", "upstream_failed"):
            st.session_state.run_state = "failed"
            st.session_state.error_message = f"DAG run ended with state: {state}"
            st.rerun()

        else:
            # Still running — auto-poll
            col_refresh, col_cancel = st.columns([1, 5])
            with col_refresh:
                if st.button("Refresh now"):
                    st.rerun()
            with col_cancel:
                if st.button("Cancel / Reset"):
                    _reset_state()
                    st.rerun()
            time.sleep(_POLL_INTERVAL_SECONDS)
            st.rerun()

    # ------------------------------------------------------------------
    # Results display
    # ------------------------------------------------------------------
    elif st.session_state.run_state == "success":
        credibility_dict = st.session_state.credibility_dict
        profile_dict = st.session_state.profile_dict
        dcf_dict = st.session_state.dcf_dict

        if not credibility_dict:
            st.error("No results available.")
            if st.button("Start new analysis"):
                _reset_state()
                st.rerun()
            st.stop()

        report = _credibility_from_dict(credibility_dict)
        profile = _profile_from_dict(profile_dict) if profile_dict else None
        dcf_result = _dcf_from_dict(dcf_dict) if dcf_dict else None

        # Apply validation layer
        from pipeline.validation_layer import ValidationLayer

        vl_result = ValidationLayer().validate(report, dcf_result=dcf_result)
        if not vl_result.passed:
            st.warning(
                f"Validation: {vl_result.error_count} error(s), "
                f"{vl_result.warning_count} warning(s). "
                "Scores and flags have been adjusted — see detail below."
            )
        report = vl_result.adjusted_report
        dcf_result = vl_result.adjusted_dcf

        # --- Header ---
        credibility_header(report)
        st.markdown("---")

        # --- Company profile metadata ---
        if profile:
            with st.expander("Company Profile", expanded=False):
                cols = st.columns(4)
                cols[0].metric("Ticker", profile.ticker or "—")
                cols[1].metric("Country", profile.country or "—")
                cols[2].metric("SIC", profile.sic_code or "—")
                cols[3].metric("Index", profile.index or "—")
                if profile.sic_description:
                    st.caption(f"Industry: {profile.sic_description}")
                if profile.latest_annual_filing:
                    f = profile.latest_annual_filing
                    st.caption(
                        f"Latest filing: {f.filing_type} filed {f.filed_date}"
                        + (f" (period: {f.period_of_report})" if f.period_of_report else "")
                    )
                if profile.source_urls:
                    st.markdown("**Data Sources**")
                    for url in profile.source_urls:
                        source_citation(url)

        # --- Validation warnings detail ---
        if vl_result.warnings:
            with st.expander(
                f"Validation detail ({vl_result.error_count} errors, "
                f"{vl_result.warning_count} warnings)",
                expanded=False,
            ):
                for w in vl_result.warnings:
                    icon = "🔴" if w.severity == "error" else "🟡"
                    st.caption(f"{icon} **{w.code}** [{w.field}]: {w.message}")

        # --- Per-factor panels ---
        st.subheader("Factor Scores")
        for fs in report.factor_scores:
            factor_panel(fs)

        # --- DCF results ---
        if dcf_result and dcf_result.mappings:
            st.markdown("---")
            st.subheader("DCF Scenario Impact Ranges")
            st.caption(
                f"Based on {dcf_result.line_item_count} DCF row labels "
                f"from {', '.join(dcf_result.sheet_names)}."
            )
            for mapping in dcf_result.mappings:
                dcf_mapping_panel(mapping)
            if dcf_result.unmapped_factors:
                st.caption(
                    "Unmapped factors (no matching DCF line): "
                    + ", ".join(dcf_result.unmapped_factors)
                )

        # --- Export ---
        st.markdown("---")
        st.subheader("Export")
        col_pdf, col_json, col_new = st.columns([1, 1, 2])

        with col_pdf:
            pdf_bytes = to_pdf(report, profile, dcf_result)
            st.download_button(
                label="Download PDF",
                data=pdf_bytes,
                file_name=f"esg_signal_{report.ticker or 'report'}.pdf",
                mime="application/pdf",
            )

        with col_json:
            json_str = to_json(report, profile, dcf_result)
            st.download_button(
                label="Download JSON",
                data=json_str,
                file_name=f"esg_signal_{report.ticker or 'report'}.json",
                mime="application/json",
            )

        with col_new:
            if st.button("New analysis"):
                _reset_state()
                st.rerun()


if __name__ == "__main__":
    main()

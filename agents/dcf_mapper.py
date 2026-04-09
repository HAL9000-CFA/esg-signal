"""
DCF Line Item Mapper — issue #13.

Maps each SASB-material ESG factor to the specific line items in an analyst's
Excel DCF model that are most likely to be affected by that risk.  Scenario
impact ranges for each mapped line are generated from published regulatory
precedents and, where available, from actual enforcement records in
data/processed/.

Privacy guarantee
-----------------
The Excel file is parsed entirely on the analyst's local machine by openpyxl.
Only the *text labels* of DCF rows are sent to Claude for mapping — no
financial figures, formulas, or other model content ever leave the local
environment.

Pipeline
--------
1. _parse_excel()       — openpyxl extracts row labels from all sheets (local)
2. _map_with_claude()   — Claude matches labels to ESG factors (text only)
3. _get_scenario_range() — Python derives low/mid/high from regulatory precedents
4. Returns DcfMapperResult

Scenario range sources
----------------------
- EPA ECHO civil penalties  : EPA ECHO Enforcement & Compliance data (2023)
  Median civil penalty ~$200 K; 95th percentile ~$15 M (oil/gas: up to $5.5 B)
- EU ETS excess emissions   : Art. 16, Directive 2003/87/EC — €100/tonne CO2
  Moderate excess (5% of verified submissions) = €50 K – €20 M
- UK EA enforcement         : Environment Agency Annual Report 2023
  Significant enforcement actions: £100 K – £100 M
- Historical impairments    : public financial statements for comparable
  stranded-asset write-downs and environmental remediation provisions

Where a company-specific regulatory CSV exists in data/processed/, actual
enforcement amounts from that file are used to anchor the mid-range; the
built-in table provides low/high bounds and covers gaps.
"""

import json
import logging
import re
from pathlib import Path
from textwrap import dedent
from typing import Dict, List, Optional, Tuple

from pipeline.llm_client import call_claude
from pipeline.models import (
    CredibilityReport,
    DcfLineItem,
    DcfMapperResult,
    EsgDcfMapping,
    MaterialFactor,
)

LOGGER = logging.getLogger(__name__)

_MODEL = "claude-opus-4-5"
_AGENT = "dcf_mapper"

# ---------------------------------------------------------------------------
# Scenario precedent table (all amounts in USD unless noted)
# Sources cited in module docstring.
# ---------------------------------------------------------------------------

_PRECEDENTS: Dict[str, Dict] = {
    # Environment
    "ghg_emissions": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 50_000_000,
        "currency": "USD",
        "source": (
            "EU ETS: €100/tonne CO2 excess (Art. 16, Dir. 2003/87/EC); "
            "EPA ECHO median civil penalty ~$200K (EPA ECHO 2023 data)"
        ),
    },
    "air_quality": {
        "low": 50_000,
        "mid": 500_000,
        "high": 5_000_000,
        "currency": "USD",
        "source": "EPA ECHO CAA civil penalties: median ~$100K, 95th pct ~$5M (EPA ECHO 2023)",
    },
    "water_management": {
        "low": 100_000,
        "mid": 1_000_000,
        "high": 15_000_000,
        "currency": "USD",
        "source": (
            "EPA ECHO CWA civil penalties: median ~$200K (EPA ECHO 2023); "
            "UK EA enforcement: £100K–£50M (EA Annual Report 2023)"
        ),
    },
    "ecological_impacts": {
        "low": 500_000,
        "mid": 10_000_000,
        "high": 200_000_000,
        "currency": "USD",
        "source": (
            "EPA NRD settlements: median ~$2M, major cases >$100M; "
            "UK EA habitat restoration orders: £500K–£50M (EA 2023)"
        ),
    },
    "waste_management": {
        "low": 50_000,
        "mid": 500_000,
        "high": 5_000_000,
        "currency": "USD",
        "source": "EPA ECHO RCRA hazardous waste penalties: median ~$150K (EPA ECHO 2023)",
    },
    "packaging": {
        "low": 50_000,
        "mid": 300_000,
        "high": 3_000_000,
        "currency": "USD",
        "source": "EU Packaging & Packaging Waste Directive compliance costs; EPA ECHO RCRA data",
    },
    "critical_incident_risk": {
        "low": 5_000_000,
        "mid": 100_000_000,
        "high": 5_000_000_000,
        "currency": "USD",
        "source": (
            "Historical precedents: BP Deepwater Horizon ~$65B total; "
            "mid-size oil spill settlements $50M–$500M (NRC incident database)"
        ),
    },
    "climate_physical_risk": {
        "low": 1_000_000,
        "mid": 20_000_000,
        "high": 500_000_000,
        "currency": "USD",
        "source": (
            "IPCC AR6 physical risk cost estimates; historical impairments: "
            "Shell NS write-down $4.5B (2022), comparable sector range 5–30% PP&E"
        ),
    },
    "energy_management": {
        "low": 100_000,
        "mid": 1_000_000,
        "high": 10_000_000,
        "currency": "USD",
        "source": "EU ETS carbon cost pass-through; EPA Clean Power Plan compliance costs",
    },
    # Social Capital
    "data_security": {
        "low": 500_000,
        "mid": 10_000_000,
        "high": 500_000_000,
        "currency": "USD",
        "source": (
            "GDPR maximum fine: 4% global turnover; average data breach cost $4.5M "
            "(IBM 2023); major cases: Meta $1.3B, Amazon $877M"
        ),
    },
    "customer_privacy": {
        "low": 200_000,
        "mid": 5_000_000,
        "high": 200_000_000,
        "currency": "USD",
        "source": (
            "GDPR Art. 83 administrative fines; FTC enforcement actions 2023; "
            "CCPA penalties $2,500–$7,500 per violation"
        ),
    },
    "product_quality_safety": {
        "low": 1_000_000,
        "mid": 20_000_000,
        "high": 1_000_000_000,
        "currency": "USD",
        "source": (
            "FDA recall costs: median $10M, major cases >$1B; "
            "EU product liability claims (RAPEX database 2023)"
        ),
    },
    "access_affordability": {
        "low": 100_000,
        "mid": 1_000_000,
        "high": 10_000_000,
        "currency": "USD",
        "source": "FTC / CMA access & affordability enforcement actions 2023",
    },
    "financial_inclusion": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 50_000_000,
        "currency": "USD",
        "source": "CFPB fair lending enforcement actions; FCA enforcement 2023",
    },
    "selling_practices": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 100_000_000,
        "currency": "USD",
        "source": "FTC / CMA consumer protection enforcement; ASA/CAP significant rulings 2023",
    },
    "community_relations": {
        "low": 1_000_000,
        "mid": 10_000_000,
        "high": 200_000_000,
        "currency": "USD",
        "source": "IFC performance standards; historical community compensation: Freeport McMoRan $600M",
    },
    # Human Capital
    "employee_engagement_diversity": {
        "low": 100_000,
        "mid": 2_000_000,
        "high": 30_000_000,
        "currency": "USD",
        "source": "EEOC enforcement actions: median settlement $300K; major cases >$100M (2023)",
    },
    "labor_practices": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 50_000_000,
        "currency": "USD",
        "source": "DOL wage and hour enforcement: median $200K; NLRB back-pay orders (2023)",
    },
    "employee_health_safety": {
        "low": 100_000,
        "mid": 2_000_000,
        "high": 20_000_000,
        "currency": "USD",
        "source": "OSHA civil penalties: max $15,625/violation; major incidents up to $20M+ (OSHA 2023)",
    },
    # Business Model & Innovation
    "product_lifecycle_management": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 50_000_000,
        "currency": "USD",
        "source": "EU WEEE / EcoDesign Regulation compliance costs; EPR scheme contributions",
    },
    "materials_sourcing": {
        "low": 200_000,
        "mid": 2_000_000,
        "high": 20_000_000,
        "currency": "USD",
        "source": "Supply chain remediation costs; EU CBAM carbon border adjustment (2026 full phase-in)",
    },
    "supply_chain_management": {
        "low": 500_000,
        "mid": 5_000_000,
        "high": 50_000_000,
        "currency": "USD",
        "source": (
            "UK Modern Slavery Act compliance; EU CSDDD (2024) supply chain due diligence; "
            "historical supply chain disruption costs"
        ),
    },
    "business_model_resilience": {
        "low": 1_000_000,
        "mid": 20_000_000,
        "high": 500_000_000,
        "currency": "USD",
        "source": "TCFD transition risk scenarios; IEA Net Zero pathway capex estimates",
    },
    # Leadership & Governance
    "systemic_risk_management": {
        "low": 1_000_000,
        "mid": 50_000_000,
        "high": 1_000_000_000,
        "currency": "USD",
        "source": (
            "Basel III / Pillar 2 capital add-on precedents; "
            "FRB/PRA systemic risk surcharges; historical bank write-downs"
        ),
    },
    "business_ethics": {
        "low": 500_000,
        "mid": 10_000_000,
        "high": 500_000_000,
        "currency": "USD",
        "source": (
            "DOJ FCPA enforcement actions: median settlement $20M (2023); "
            "SFO / FCA enforcement: major cases Rolls-Royce $800M, Airbus $4B"
        ),
    },
    "competitive_behavior": {
        "low": 1_000_000,
        "mid": 20_000_000,
        "high": 2_000_000_000,
        "currency": "USD",
        "source": (
            "EU competition fines: up to 10% global turnover; "
            "DOJ antitrust: Google $5B, record EU fines"
        ),
    },
}

# Default for any factor_id not in the table (uses generic ranges)
_DEFAULT_PRECEDENT = {
    "low": 500_000,
    "mid": 5_000_000,
    "high": 50_000_000,
    "currency": "USD",
    "source": "Generic ESG enforcement precedent range — no factor-specific data available",
}

# ---------------------------------------------------------------------------
# Claude prompt
# ---------------------------------------------------------------------------

_MAPPING_SYSTEM = """You are a financial analyst mapping ESG risks to DCF model line items.

You will receive:
  1. A list of line item labels parsed from an Excel DCF model, grouped by sheet.
  2. A list of material ESG factors with their SASB financial impact categories.

Task: For each ESG factor, identify which DCF line items are most likely to be
affected by that risk, based on the SASB financial impact category.

Rules:
- Only use labels EXACTLY as they appear in the list — do not rephrase, abbreviate, or invent new ones
- Only map a factor if at least one line item is clearly relevant to that financial impact category
- A cost_impact factor maps to cost/expense/opex lines
- A revenue_impact factor maps to revenue/sales/income lines
- An asset_impact factor maps to asset/capex/property/goodwill lines
- A liability_impact factor maps to liability/provision/penalty/legal lines
- Return a JSON object: {"factor_id": ["label1", "label2"], ...}
- Omit factors where no clear match exists
- Return valid JSON only, no prose, no markdown fences"""

# ---------------------------------------------------------------------------
# Excel parsing helpers
# ---------------------------------------------------------------------------

# Year-like numbers to skip (not financial line item labels)
_YEAR_RE = re.compile(r"^\s*(19|20)\d{2}[Ee]?\s*$")
# Labels that are purely numeric should be skipped
_NUMERIC_RE = re.compile(r"^\s*[-+]?[\d,.\s]+%?\s*$")


def _is_valid_label(val) -> bool:
    """Return True if the cell value looks like a financial line item label."""
    if not isinstance(val, str):
        return False
    stripped = val.strip()
    if len(stripped) < 3:
        return False
    if _YEAR_RE.match(stripped):
        return False
    if _NUMERIC_RE.match(stripped):
        return False
    return True


# ---------------------------------------------------------------------------
# DcfMapper
# ---------------------------------------------------------------------------


class DcfMapper:
    """
    Maps material ESG factors to DCF line items from an analyst's Excel model.

    Usage::

        mapper = DcfMapper()
        result = mapper.map(
            excel_path="path/to/model.xlsx",
            material_factors=relevance_result.material_factors,
            credibility_report=credibility_report,     # optional
            regulatory_paths=data_gatherer_result.regulatory_paths,  # optional
            run_id="airflow_run_123",
        )
    """

    def map(
        self,
        excel_path: str,
        material_factors: List[MaterialFactor],
        credibility_report: Optional[CredibilityReport] = None,
        regulatory_paths: Optional[Dict[str, str]] = None,
        run_id: Optional[str] = None,
    ) -> DcfMapperResult:
        """
        Map ESG factors to DCF line items and generate scenario ranges.

        The Excel file is never transmitted anywhere. Only text row labels are
        sent to Claude for mapping — no financial figures leave the local machine.

        Args:
            excel_path:         Local path to the analyst's Excel DCF model.
            material_factors:   Material factors from RelevanceFilter.
            credibility_report: Optional CredibilityReport for attaching flags.
            regulatory_paths:   Dict of source → processed CSV path for actual
                                enforcement data (used to refine scenario ranges).
            run_id:             Airflow run ID forwarded to the audit log.

        Returns:
            DcfMapperResult with per-factor line item mappings and scenario ranges.
        """
        regulatory_paths = regulatory_paths or {}
        errors: List[str] = []

        # Step 1: parse Excel locally — extract row labels only
        line_items, sheet_names, parse_errors = self._parse_excel(excel_path)
        errors.extend(parse_errors)

        if not line_items:
            return DcfMapperResult(
                ticker=None,
                excel_path=excel_path,
                sheet_names=sheet_names,
                line_item_count=0,
                mappings=[],
                unmapped_factors=[f.name for f in material_factors],
                errors=errors or ["No line items found in workbook"],
            )

        # Step 2: Claude maps label strings to factors
        raw_mapping, map_errors = self._map_with_claude(
            line_items=line_items,
            material_factors=material_factors,
            run_id=run_id,
        )
        errors.extend(map_errors)

        # Step 3: resolve labels → DcfLineItem objects; compute scenario ranges
        # Build lookup: label → first DcfLineItem with that label (case-insensitive)
        label_lookup: Dict[str, DcfLineItem] = {}
        for item in line_items:
            label_lookup.setdefault(item.label, item)

        flag_map: Dict[str, str] = {}
        if credibility_report:
            flag_map = {fs.factor_id: fs.flag for fs in credibility_report.factor_scores}

        mappings: List[EsgDcfMapping] = []
        unmapped: List[str] = []

        for factor in material_factors:
            matched_labels = raw_mapping.get(factor.factor_id, [])
            matched_items = [label_lookup[lbl] for lbl in matched_labels if lbl in label_lookup]

            if not matched_items:
                unmapped.append(factor.name)
                continue

            low, mid, high, currency, source = self._get_scenario_range(
                factor=factor, regulatory_paths=regulatory_paths
            )

            mappings.append(
                EsgDcfMapping(
                    factor_id=factor.factor_id,
                    factor_name=factor.name,
                    financial_impacts=factor.financial_impacts,
                    mapped_line_items=matched_items,
                    scenario_low=low,
                    scenario_mid=mid,
                    scenario_high=high,
                    scenario_currency=currency,
                    scenario_source=source,
                    credibility_flag=flag_map.get(factor.factor_id),
                )
            )

        return DcfMapperResult(
            ticker=credibility_report.ticker if credibility_report else None,
            excel_path=excel_path,
            sheet_names=sheet_names,
            line_item_count=len(line_items),
            mappings=mappings,
            unmapped_factors=unmapped,
            errors=errors,
        )

    # ------------------------------------------------------------------
    # Step 1 — Excel parsing (local, nothing transmitted)
    # ------------------------------------------------------------------

    def _parse_excel(self, excel_path: str) -> Tuple[List[DcfLineItem], List[str], List[str]]:
        """
        Parse a local Excel workbook and return (line_items, sheet_names, errors).

        Scans every sheet, every row, checking the first three columns for text
        labels that look like DCF line items.  Numbers, years, and empty cells
        are skipped.  Only the label text is retained — no values are read.
        """
        errors: List[str] = []
        line_items: List[DcfLineItem] = []
        sheet_names: List[str] = []

        try:
            import openpyxl
        except ImportError:
            return [], [], ["openpyxl not installed"]

        if not Path(excel_path).exists():
            return [], [], [f"Excel file not found: {excel_path}"]

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            return [], [], [f"Could not open workbook: {exc}"]

        try:
            for sheet_name in wb.sheetnames:
                sheet_names.append(sheet_name)
                try:
                    ws = wb[sheet_name]
                    for row_idx, row in enumerate(
                        ws.iter_rows(max_row=300, values_only=True), start=1
                    ):
                        # Check first 3 columns for a label
                        for col_idx in range(min(3, len(row))):
                            cell_val = row[col_idx]
                            if _is_valid_label(cell_val):
                                label = cell_val.strip()
                                line_items.append(
                                    DcfLineItem(
                                        sheet_name=sheet_name,
                                        row_index=row_idx,
                                        label=label,
                                    )
                                )
                                break  # only one label per row
                except Exception as exc:
                    errors.append(f"Error reading sheet '{sheet_name}': {exc}")
        finally:
            wb.close()

        # Deduplicate: keep first occurrence of each label
        seen: set = set()
        deduped: List[DcfLineItem] = []
        for item in line_items:
            if item.label not in seen:
                seen.add(item.label)
                deduped.append(item)

        return deduped, sheet_names, errors

    # ------------------------------------------------------------------
    # Step 2 — Claude mapping (labels only, no values sent)
    # ------------------------------------------------------------------

    def _map_with_claude(
        self,
        line_items: List[DcfLineItem],
        material_factors: List[MaterialFactor],
        run_id: Optional[str],
    ) -> Tuple[Dict[str, List[str]], List[str]]:
        """
        Ask Claude to match DCF line item labels to ESG factors.

        Returns (mapping_dict, errors).
        mapping_dict: {factor_id: [label1, label2, ...]}
        """
        if not line_items or not material_factors:
            return {}, []

        # Group labels by sheet for a readable prompt
        by_sheet: Dict[str, List[str]] = {}
        for item in line_items:
            by_sheet.setdefault(item.sheet_name, []).append(item.label)

        labels_section = "\n".join(
            f'Sheet "{sheet}":\n' + "\n".join(f"  - {lbl}" for lbl in lbls)
            for sheet, lbls in by_sheet.items()
        )

        def _factor_line(f: MaterialFactor) -> str:
            line = f"- {f.factor_id}: \"{f.name}\" → affects {', '.join(f.financial_impacts)}"
            if f.description:
                line += f"\n  Context: {f.description.strip()}"
            return line

        factors_section = "\n".join(_factor_line(f) for f in material_factors)

        prompt = dedent(
            f"""
            DCF model line items (labels only — no values):
            {labels_section}

            Material ESG factors (use exact factor_id as JSON keys):
            {factors_section}

            Map each factor to the DCF line items most likely to be affected.
            Return only factors where at least one line item is a clear match.
        """
        ).strip()

        try:
            raw = call_claude(
                agent=_AGENT,
                model=_MODEL,
                version=_MODEL,
                purpose="DCF line item mapping",
                system=_MAPPING_SYSTEM,
                prompt=prompt,
                max_tokens=1024,
                temperature=0.0,
                run_id=run_id,
            )
        except Exception as exc:
            LOGGER.warning("Claude DCF mapping failed: %s", exc)
            return {}, [f"Claude mapping failed: {exc}"]

        try:
            parsed = json.loads(raw.strip())
            if not isinstance(parsed, dict):
                raise ValueError("Expected a JSON object")
        except (json.JSONDecodeError, ValueError) as exc:
            LOGGER.warning("Could not parse DCF mapping response: %s", exc)
            return {}, [f"Could not parse mapping response: {exc}"]

        # Validate: only keep labels that actually exist in the parsed set
        valid_labels = {item.label for item in line_items}
        cleaned: Dict[str, List[str]] = {}
        for factor_id, labels in parsed.items():
            if not isinstance(labels, list):
                continue
            valid = [lbl for lbl in labels if isinstance(lbl, str) and lbl in valid_labels]
            if valid:
                cleaned[factor_id] = valid

        return cleaned, []

    # ------------------------------------------------------------------
    # Step 3 — Scenario ranges (pure Python, no LLM)
    # ------------------------------------------------------------------

    def _get_scenario_range(
        self,
        factor: MaterialFactor,
        regulatory_paths: Dict[str, str],
    ) -> Tuple[Optional[float], Optional[float], Optional[float], str, str]:
        """
        Return (low, mid, high, currency, source) for a factor's scenario range.

        Priority:
        1. Actual enforcement amounts from company-specific regulatory CSVs
           (ECHO, EU ETS, EA Pollution Inventory) where available.
        2. Built-in published regulatory precedent table.

        All arithmetic is pure Python — no LLM involvement.
        """
        precedent = _PRECEDENTS.get(factor.factor_id, _DEFAULT_PRECEDENT)
        low = float(precedent["low"])
        mid = float(precedent["mid"])
        high = float(precedent["high"])
        currency = precedent["currency"]
        source = precedent["source"]

        # Attempt to refine mid from actual regulatory enforcement data
        actual_mid = self._extract_actual_mid(factor, regulatory_paths)
        if actual_mid is not None:
            mid = actual_mid
            source = source + " (mid adjusted to actual enforcement data from data/processed/)"

        return low, mid, high, currency, source

    def _extract_actual_mid(
        self, factor: MaterialFactor, regulatory_paths: Dict[str, str]
    ) -> Optional[float]:
        """
        Attempt to extract a mid-range estimate from actual regulatory CSV data.

        Looks for penalty / fine amount columns in the available CSVs and
        returns the median as the mid-range anchor.  Returns None if no
        suitable data is found.
        """
        try:
            import pandas as pd
        except ImportError:
            return None

        # Map factor dimensions / IDs to the regulatory sources most likely to
        # contain enforcement amounts for that factor.
        source_priority = self._regulatory_sources_for_factor(factor)

        for source_key in source_priority:
            path = regulatory_paths.get(source_key)
            if not path or not Path(path).exists():
                continue
            try:
                df = pd.read_csv(path)
                amounts = self._extract_penalty_amounts(df)
                if amounts and len(amounts) >= 3:
                    return float(sorted(amounts)[len(amounts) // 2])  # median
            except Exception as exc:
                LOGGER.debug("Could not read %s for scenario range: %s", path, exc)

        return None

    @staticmethod
    def _regulatory_sources_for_factor(factor: MaterialFactor) -> List[str]:
        """Return regulatory source keys in priority order for a given factor."""
        _mapping = {
            "ghg_emissions": ["ghgrp", "eu_ets"],
            "air_quality": ["echo", "ea_pollution"],
            "water_management": ["echo", "ea_pollution"],
            "ecological_impacts": ["echo", "ea_pollution"],
            "waste_management": ["echo"],
            "critical_incident_risk": ["nrc", "echo"],
            "climate_physical_risk": ["eu_ets"],
            "energy_management": ["ghgrp", "eu_ets"],
        }
        return _mapping.get(factor.factor_id, ["echo", "ea_pollution"])

    @staticmethod
    def _extract_penalty_amounts(df) -> List[float]:
        """
        Scan a regulatory DataFrame for penalty/fine amount columns and
        return all non-zero positive numeric values found.
        """
        amounts: List[float] = []
        for col in df.columns:
            col_lower = col.lower()
            if any(kw in col_lower for kw in ("penalty", "fine", "enforcement", "amount")):
                try:
                    import pandas as pd

                    vals = pd.to_numeric(df[col], errors="coerce").dropna()
                    amounts.extend(float(v) for v in vals if v > 0)
                except Exception:
                    pass
        return amounts

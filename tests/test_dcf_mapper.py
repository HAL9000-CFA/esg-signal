"""
Unit tests for agents/dcf_mapper.py.
No live API calls — Claude and openpyxl are mocked/used with real temp files.
"""

import json
from unittest.mock import patch

import openpyxl

from agents.dcf_mapper import _PRECEDENTS, DcfMapper, _is_valid_label
from pipeline.models import (
    CredibilityReport,
    DcfLineItem,
    DcfMapperResult,
    FactorScore,
    MaterialFactor,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_factor(factor_id="ghg_emissions", dimension="Environment", impacts=None):
    return MaterialFactor(
        factor_id=factor_id,
        name=(
            "GHG Emissions" if factor_id == "ghg_emissions" else factor_id.replace("_", " ").title()
        ),
        dimension=dimension,
        financial_impacts=impacts or ["cost_impact", "asset_impact"],
    )


def _make_credibility_report(factor_id="ghg_emissions", flag="amber"):
    fs = FactorScore(
        factor_id=factor_id,
        factor_name="GHG Emissions",
        score=0.55,
        flag=flag,
        stream_scores={},
        evidence=[],
        sources=[],
        narrative="",
    )
    return CredibilityReport(
        ticker="TEST",
        company_name="Test Corp",
        sasb_industry="Oil & Gas",
        factor_scores=[fs],
        overall_score=0.55,
        overall_flag=flag,
        errors=[],
    )


def _make_excel(tmp_path, rows_by_sheet=None) -> str:
    """Create a minimal Excel workbook with given rows and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Model"

    default_rows = [
        ("Revenue",),
        ("Cost of Goods Sold",),
        ("Gross Profit",),
        ("EBITDA",),
        ("Depreciation & Amortisation",),
        ("EBIT",),
        ("Interest Expense",),
        ("Net Income",),
        ("Capital Expenditure",),
        ("Free Cash Flow",),
        ("Environmental Provisions",),
        ("Remediation Liability",),
    ]

    rows = rows_by_sheet or {"DCF Model": default_rows}

    ws.title = list(rows.keys())[0]
    for row_data in list(rows.values())[0]:
        ws.append(row_data)

    # Add extra sheets if present
    for i, (sheet_name, sheet_rows) in enumerate(list(rows.items())[1:]):
        ws2 = wb.create_sheet(title=sheet_name)
        for row_data in sheet_rows:
            ws2.append(row_data)

    path = str(tmp_path / "test_dcf.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# _is_valid_label
# ---------------------------------------------------------------------------


class TestIsValidLabel:
    def test_normal_label_is_valid(self):
        assert _is_valid_label("Capital Expenditure") is True

    def test_short_label_rejected(self):
        assert _is_valid_label("ab") is False

    def test_year_rejected(self):
        assert _is_valid_label("2025") is False
        assert _is_valid_label("2030") is False

    def test_numeric_string_rejected(self):
        assert _is_valid_label("1,234,567") is False
        assert _is_valid_label("12.5%") is False

    def test_non_string_rejected(self):
        assert _is_valid_label(12345) is False
        assert _is_valid_label(None) is False

    def test_empty_string_rejected(self):
        assert _is_valid_label("") is False

    def test_whitespace_only_rejected(self):
        assert _is_valid_label("   ") is False

    def test_label_with_special_chars_valid(self):
        assert _is_valid_label("EBIT(DA)") is True


# ---------------------------------------------------------------------------
# DcfMapper._parse_excel
# ---------------------------------------------------------------------------


class TestParseExcel:
    def test_parses_labels_from_workbook(self, tmp_path):
        path = _make_excel(tmp_path)
        mapper = DcfMapper()
        items, sheets, errors = mapper._parse_excel(path)

        assert errors == []
        labels = [i.label for i in items]
        assert "Revenue" in labels
        assert "Capital Expenditure" in labels
        assert "EBITDA" in labels

    def test_missing_file_returns_error(self):
        mapper = DcfMapper()
        items, sheets, errors = mapper._parse_excel("/nonexistent/path/model.xlsx")
        assert items == []
        assert any("not found" in e for e in errors)

    def test_sheet_name_recorded_on_items(self, tmp_path):
        path = _make_excel(tmp_path)
        mapper = DcfMapper()
        items, sheets, errors = mapper._parse_excel(path)

        assert "DCF Model" in sheets
        assert all(i.sheet_name == "DCF Model" for i in items)

    def test_row_index_recorded(self, tmp_path):
        path = _make_excel(tmp_path)
        mapper = DcfMapper()
        items, _, _ = mapper._parse_excel(path)

        revenue_item = next(i for i in items if i.label == "Revenue")
        assert revenue_item.row_index == 1

    def test_deduplication(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(("Revenue",))
        ws.append(("Revenue",))  # duplicate
        ws.append(("Cost",))
        path = str(tmp_path / "dup.xlsx")
        wb.save(path)

        mapper = DcfMapper()
        items, _, _ = mapper._parse_excel(path)
        labels = [i.label for i in items]
        assert labels.count("Revenue") == 1

    def test_year_columns_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(("Revenue", 2024, 2025, 2026))  # years in later cols
        ws.append(("2024",))  # year as row label — should be skipped
        path = str(tmp_path / "years.xlsx")
        wb.save(path)

        mapper = DcfMapper()
        items, _, _ = mapper._parse_excel(path)
        labels = [i.label for i in items]
        assert "Revenue" in labels
        assert "2024" not in labels


# ---------------------------------------------------------------------------
# DcfMapper._map_with_claude
# ---------------------------------------------------------------------------


class TestMapWithClaude:
    def _items(self):
        return [
            DcfLineItem("DCF", 1, "Revenue"),
            DcfLineItem("DCF", 2, "Capital Expenditure"),
            DcfLineItem("DCF", 3, "Environmental Provisions"),
        ]

    def test_returns_dict_of_lists(self):
        factors = [_make_factor()]
        response = json.dumps(
            {"ghg_emissions": ["Capital Expenditure", "Environmental Provisions"]}
        )

        with patch("agents.dcf_mapper.call_claude", return_value=response):
            mapping, errors = DcfMapper()._map_with_claude(self._items(), factors, None)

        assert errors == []
        assert mapping["ghg_emissions"] == ["Capital Expenditure", "Environmental Provisions"]

    def test_invalid_labels_filtered_out(self):
        factors = [_make_factor()]
        # Claude returns a label that doesn't exist in the parsed set
        response = json.dumps({"ghg_emissions": ["Capital Expenditure", "INVENTED_LINE"]})

        with patch("agents.dcf_mapper.call_claude", return_value=response):
            mapping, errors = DcfMapper()._map_with_claude(self._items(), factors, None)

        assert "INVENTED_LINE" not in mapping.get("ghg_emissions", [])
        assert "Capital Expenditure" in mapping.get("ghg_emissions", [])

    def test_claude_failure_returns_empty_with_error(self):
        factors = [_make_factor()]
        with patch("agents.dcf_mapper.call_claude", side_effect=Exception("API down")):
            mapping, errors = DcfMapper()._map_with_claude(self._items(), factors, None)

        assert mapping == {}
        assert any("failed" in e.lower() for e in errors)

    def test_invalid_json_returns_error(self):
        factors = [_make_factor()]
        with patch("agents.dcf_mapper.call_claude", return_value="not json"):
            mapping, errors = DcfMapper()._map_with_claude(self._items(), factors, None)

        assert mapping == {}
        assert errors

    def test_empty_line_items_returns_empty(self):
        factors = [_make_factor()]
        mapping, errors = DcfMapper()._map_with_claude([], factors, None)
        assert mapping == {}
        assert errors == []


# ---------------------------------------------------------------------------
# DcfMapper._get_scenario_range
# ---------------------------------------------------------------------------


class TestGetScenarioRange:
    def test_known_factor_returns_precedent_values(self):
        factor = _make_factor("ghg_emissions")
        low, mid, high, currency, source = DcfMapper()._get_scenario_range(factor, {})

        assert low == float(_PRECEDENTS["ghg_emissions"]["low"])
        assert high == float(_PRECEDENTS["ghg_emissions"]["high"])
        assert currency == "USD"
        assert "EU ETS" in source

    def test_unknown_factor_returns_default(self):
        factor = _make_factor("unknown_factor_xyz")
        low, mid, high, currency, source = DcfMapper()._get_scenario_range(factor, {})

        assert low > 0
        assert mid > 0
        assert high > 0
        assert "Generic" in source

    def test_low_less_than_mid_less_than_high(self):
        for factor_id in _PRECEDENTS:
            factor = _make_factor(factor_id)
            low, mid, high, _, _ = DcfMapper()._get_scenario_range(factor, {})
            assert low < mid < high, f"Failed for {factor_id}: {low} < {mid} < {high}"

    def test_actual_data_adjusts_mid(self, tmp_path):
        import pandas as pd

        csv_path = tmp_path / "echo_TEST.csv"
        pd.DataFrame({"PENALTY_AMOUNT": [100_000, 200_000, 300_000, 400_000, 500_000]}).to_csv(
            csv_path, index=False
        )
        factor = _make_factor("air_quality")
        _, mid, _, _, source = DcfMapper()._get_scenario_range(factor, {"echo": str(csv_path)})

        # median of [100K..500K] = 300K
        assert mid == 300_000.0
        assert "actual enforcement data" in source


# ---------------------------------------------------------------------------
# DcfMapper.map — integration (Claude mocked, real Excel)
# ---------------------------------------------------------------------------


class TestDcfMapperMap:
    def _run(self, tmp_path, factors=None, credibility_report=None, claude_response=None):
        path = _make_excel(tmp_path)
        factors = factors or [_make_factor()]
        mapping_response = claude_response or json.dumps(
            {"ghg_emissions": ["Capital Expenditure", "Environmental Provisions"]}
        )

        with patch("agents.dcf_mapper.call_claude", return_value=mapping_response):
            return DcfMapper().map(
                excel_path=path,
                material_factors=factors,
                credibility_report=credibility_report,
            )

    def test_returns_dcf_mapper_result(self, tmp_path):
        result = self._run(tmp_path)
        assert isinstance(result, DcfMapperResult)

    def test_mapped_factors_appear_in_mappings(self, tmp_path):
        result = self._run(tmp_path)
        factor_ids = {m.factor_id for m in result.mappings}
        assert "ghg_emissions" in factor_ids

    def test_mapped_line_items_populated(self, tmp_path):
        result = self._run(tmp_path)
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        labels = [li.label for li in mapping.mapped_line_items]
        assert "Capital Expenditure" in labels

    def test_scenario_ranges_set(self, tmp_path):
        result = self._run(tmp_path)
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        assert mapping.scenario_low is not None
        assert mapping.scenario_mid is not None
        assert mapping.scenario_high is not None
        assert mapping.scenario_low < mapping.scenario_mid < mapping.scenario_high

    def test_credibility_flag_attached_when_report_provided(self, tmp_path):
        report = _make_credibility_report(flag="red")
        result = self._run(tmp_path, credibility_report=report)
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        assert mapping.credibility_flag == "red"

    def test_credibility_flag_none_without_report(self, tmp_path):
        result = self._run(tmp_path, credibility_report=None)
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        assert mapping.credibility_flag is None

    def test_unmapped_factor_listed(self, tmp_path):
        # Claude returns no mapping for data_security
        factors = [_make_factor("ghg_emissions"), _make_factor("data_security", "Social Capital")]
        result = self._run(
            tmp_path,
            factors=factors,
            claude_response=json.dumps({"ghg_emissions": ["Capital Expenditure"]}),
        )
        assert "Data Security" in result.unmapped_factors

    def test_missing_excel_file_produces_error(self):
        factor = _make_factor()
        with patch("agents.dcf_mapper.call_claude", return_value="{}"):
            result = DcfMapper().map(
                excel_path="/nonexistent/model.xlsx",
                material_factors=[factor],
            )
        assert result.errors

    def test_excel_path_recorded_on_result(self, tmp_path):
        path = _make_excel(tmp_path)
        with patch("agents.dcf_mapper.call_claude", return_value="{}"):
            result = DcfMapper().map(excel_path=path, material_factors=[_make_factor()])
        assert result.excel_path == path

    def test_line_item_count_matches_parsed(self, tmp_path):
        path = _make_excel(tmp_path)
        with patch("agents.dcf_mapper.call_claude", return_value="{}"):
            result = DcfMapper().map(excel_path=path, material_factors=[_make_factor()])
        assert result.line_item_count > 0

    def test_sheet_names_recorded(self, tmp_path):
        path = _make_excel(tmp_path)
        with patch("agents.dcf_mapper.call_claude", return_value="{}"):
            result = DcfMapper().map(excel_path=path, material_factors=[_make_factor()])
        assert "DCF Model" in result.sheet_names

    def test_ticker_from_credibility_report(self, tmp_path):
        report = _make_credibility_report()
        result = self._run(tmp_path, credibility_report=report)
        assert result.ticker == "TEST"

    def test_financial_impacts_propagated(self, tmp_path):
        factor = _make_factor("ghg_emissions", impacts=["cost_impact", "liability_impact"])
        result = self._run(tmp_path, factors=[factor])
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        assert "cost_impact" in mapping.financial_impacts
        assert "liability_impact" in mapping.financial_impacts

    def test_multi_sheet_workbook(self, tmp_path):
        rows = {
            "Assumptions": [("WACC",), ("Terminal Growth Rate",), ("Carbon Price",)],
            "DCF Model": [("Revenue",), ("EBIT",), ("Capital Expenditure",)],
        }
        path = _make_excel(tmp_path, rows_by_sheet=rows)
        response = json.dumps({"ghg_emissions": ["Capital Expenditure", "Carbon Price"]})

        with patch("agents.dcf_mapper.call_claude", return_value=response):
            result = DcfMapper().map(
                excel_path=path,
                material_factors=[_make_factor()],
            )

        assert len(result.sheet_names) == 2
        mapping = next(m for m in result.mappings if m.factor_id == "ghg_emissions")
        labels = [li.label for li in mapping.mapped_line_items]
        assert "Carbon Price" in labels or "Capital Expenditure" in labels

    def test_scenario_currency_is_usd(self, tmp_path):
        result = self._run(tmp_path)
        for m in result.mappings:
            assert m.scenario_currency == "USD"

    def test_scenario_source_cites_real_precedent(self, tmp_path):
        result = self._run(tmp_path)
        for m in result.mappings:
            assert m.scenario_source  # not empty


# ---------------------------------------------------------------------------
# Precedent table integrity
# ---------------------------------------------------------------------------


class TestPrecedentTable:
    def test_all_precedents_have_required_keys(self):
        required = {"low", "mid", "high", "currency", "source"}
        for factor_id, data in _PRECEDENTS.items():
            missing = required - data.keys()
            assert not missing, f"{factor_id} missing keys: {missing}"

    def test_all_precedents_low_less_than_high(self):
        for factor_id, data in _PRECEDENTS.items():
            assert data["low"] < data["high"], f"{factor_id}: low >= high"

    def test_all_precedents_have_non_empty_source(self):
        for factor_id, data in _PRECEDENTS.items():
            assert data["source"], f"{factor_id} has empty source"

    def test_all_precedents_positive_values(self):
        for factor_id, data in _PRECEDENTS.items():
            for key in ("low", "mid", "high"):
                assert data[key] > 0, f"{factor_id}.{key} is not positive"
